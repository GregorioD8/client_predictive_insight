# data_ops.py
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule

# Function to add headers with styling
def add_headers(ws):
    headers = ["Session ID", "Client ID", "Session Date", "Mood Score", "Stress Level", "AI Recommendations"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# Function to add data and apply conditional formatting
def add_data(ws, data):
    for row in data:
        ws.append(row)
    
    # Apply conditional formatting to highlight high stress levels
    stress_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add("E2:E100", CellIsRule(operator="greaterThan", formula=["6"], fill=stress_fill))